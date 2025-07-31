import os
import requests
from pathlib import Path
import mimetypes
from typing import Dict, Any, Optional, Tuple, List
import streamlit as st
from urllib.parse import urlparse, unquote
import cloudinary
import cloudinary.uploader
from urllib.parse import urlparse
import time
import hashlib

# Document processing libraries
import PyPDF2
import docx
import openpyxl
from bs4 import BeautifulSoup
import pandas as pd

class DocumentProcessor:
    def __init__(self, downloads_dir: str, cloudinary_url: str):
        self.downloads_dir = Path(downloads_dir)
        self.downloads_dir.mkdir(exist_ok=True)
        
        # Configure Cloudinary
        if cloudinary_url:
            cloudinary.config(cloudinary_url=cloudinary_url)
            st.success("Cloudinary configured successfully")
        else:
            st.warning("Cloudinary URL not configured - files will only be stored locally")
            
    def validate_url(self, url: str) -> Tuple[bool, Optional[int], Optional[str]]:
        """Validate URL accessibility before attempting download"""
        try:
            # Make a HEAD request to check if URL is accessible
            response = requests.head(url, timeout=10, allow_redirects=True)
            
            # Consider 2xx and 3xx status codes as valid
            if 200 <= response.status_code < 400:
                return True, response.status_code, None
            else:
                return False, response.status_code, f"HTTP {response.status_code}"
                
        except requests.exceptions.Timeout:
            return False, None, "Request timeout"
        except requests.exceptions.ConnectionError:
            return False, None, "Connection error"
        except requests.exceptions.RequestException as e:
            return False, None, f"Request error: {str(e)}"
        except Exception as e:
            return False, None, f"Unexpected error: {str(e)}"
    
    def download_document(self, url: str, country: str, visa_type: str) -> Optional[Dict[str, Any]]:
        """Download document from URL and upload to Cloudinary"""
        try:
            # Validate URL first
            is_valid, status_code, error_msg = self.validate_url(url)
            if not is_valid:
                st.error(f"Cannot download from invalid URL: {url} (Status: {status_code}, Error: {error_msg})")
                return None
            
            # Make the actual download request
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
            }
            
            response = requests.get(url, headers=headers, timeout=30, stream=True)
            response.raise_for_status()
            
            # Determine file format and extension
            content_type = response.headers.get('content-type', '').lower()
            file_format, file_extension = self._determine_file_format(url, content_type)
            
            # Generate filename
            filename = self._generate_filename(url, country, visa_type, file_extension)
            file_path = self.downloads_dir / filename
            
            # Download and save file
            with open(file_path, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    f.write(chunk)
            
            file_size = file_path.stat().st_size
            
            # Upload to Cloudinary
            cloudinary_url = self._upload_to_cloudinary(file_path, filename)
            
            file_info = {
                'filename': filename,
                'file_path': str(file_path),
                'file_format': file_format,
                'file_size_bytes': file_size,
                'cloudinary_url': cloudinary_url,
                'download_url': url
            }
            
            st.success(f"Downloaded: {filename} ({file_size:,} bytes)")
            if cloudinary_url:
                st.info(f"Uploaded to Cloudinary: {cloudinary_url}")
            
            return file_info
            
        except requests.exceptions.RequestException as e:
            st.error(f"Error downloading {url}: {e}")
            return None
        except Exception as e:
            st.error(f"Unexpected error downloading {url}: {e}")
            return None
    
    def _determine_file_format(self, url: str, content_type: str) -> Tuple[str, str]:
        """Determine file format and appropriate extension"""
        url_lower = url.lower()
        
        # Check URL extension first
        if url_lower.endswith('.pdf'):
            return 'PDF', '.pdf'
        elif url_lower.endswith(('.docx', '.doc')):
            return 'DOCX' if url_lower.endswith('.docx') else 'DOC', '.docx' if url_lower.endswith('.docx') else '.doc'
        elif url_lower.endswith(('.xlsx', '.xls')):
            return 'XLSX' if url_lower.endswith('.xlsx') else 'XLS', '.xlsx' if url_lower.endswith('.xlsx') else '.xls'
        elif url_lower.endswith(('.pptx', '.ppt')):
            return 'PPTX' if url_lower.endswith('.pptx') else 'PPT', '.pptx' if url_lower.endswith('.pptx') else '.ppt'
        elif url_lower.endswith('.txt'):
            return 'TXT', '.txt'
        
        # Check content type
        if 'pdf' in content_type:
            return 'PDF', '.pdf'
        elif 'word' in content_type or 'officedocument.wordprocessingml' in content_type:
            return 'DOCX', '.docx'
        elif 'excel' in content_type or 'officedocument.spreadsheetml' in content_type:
            return 'XLSX', '.xlsx'
        elif 'powerpoint' in content_type or 'officedocument.presentationml' in content_type:
            return 'PPTX', '.pptx'
        elif 'text/plain' in content_type:
            return 'TXT', '.txt'
        elif 'html' in content_type:
            return 'HTML', '.html'
        
        # Default to HTML for web pages
        return 'HTML', '.html'
    
    def _generate_filename(self, url: str, country: str, visa_type: str, extension: str) -> str:
        """Generate a unique filename for the downloaded document"""
        # Parse URL to get a base name
        parsed_url = urlparse(url)
        path_parts = parsed_url.path.strip('/').split('/')
        
        # Try to get a meaningful name from the URL
        if path_parts and path_parts[-1]:
            base_name = unquote(path_parts[-1])
            # Remove existing extension if present
            if '.' in base_name:
                base_name = '.'.join(base_name.split('.')[:-1])
        else:
            base_name = parsed_url.netloc.replace('.', '_')
        
        # Clean the base name
        base_name = ''.join(c for c in base_name if c.isalnum() or c in '-_')[:50]
        
        # Create filename with country and visa type
        country_clean = ''.join(c for c in country if c.isalnum())[:10]
        visa_clean = ''.join(c for c in visa_type.replace(' ', '_') if c.isalnum() or c == '_')[:15]
        
        # Add timestamp to ensure uniqueness
        timestamp = str(int(time.time()))[-6:]  # Last 6 digits of timestamp
        
        filename = f"{country_clean}_{visa_clean}_{base_name}_{timestamp}{extension}"
        
        return filename
    
    def _upload_to_cloudinary(self, file_path: Path, filename: str) -> Optional[str]:
        """Upload file to Cloudinary and return URL"""
        try:
            if not hasattr(cloudinary.config(), 'cloud_name') or not cloudinary.config().cloud_name:
                return None
            
            # Upload to Cloudinary
            result = cloudinary.uploader.upload(
                str(file_path),
                public_id=f"immigration_docs/{filename}",
                resource_type="auto",
                use_filename=True,
                unique_filename=False
            )
            
            return result.get('secure_url')
            
        except Exception as e:
            st.warning(f"Failed to upload to Cloudinary: {e}")
            return None
    
    def extract_text(self, file_path: str) -> str:
        """Extract text content from various file formats"""
        file_path = Path(file_path)
        
        if not file_path.exists():
            st.error(f"File not found: {file_path}")
            return ""
        
        file_extension = file_path.suffix.lower()
        
        try:
            if file_extension == '.pdf':
                return self._extract_pdf_text(file_path)
            elif file_extension in ['.docx', '.doc']:
                return self._extract_word_text(file_path)
            elif file_extension in ['.xlsx', '.xls']:
                return self._extract_excel_text(file_path)
            elif file_extension == '.txt':
                return self._extract_txt_text(file_path)
            elif file_extension in ['.html', '.htm']:
                return self._extract_html_text(file_path)
            else:
                st.warning(f"Unsupported file format: {file_extension}")
                return ""
                
        except Exception as e:
            st.error(f"Error extracting text from {file_path}: {e}")
            return ""
    
    def _extract_pdf_text(self, file_path: Path) -> str:
        """Extract text from PDF files"""
        text = ""
        try:
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                for page in pdf_reader.pages:
                    text += page.extract_text() + "\n"
        except Exception as e:
            st.error(f"Error reading PDF {file_path}: {e}")
        
        return text.strip()
    
    def _extract_word_text(self, file_path: Path) -> str:
        """Extract text from Word documents"""
        try:
            doc = docx.Document(file_path)
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            return text.strip()
        except Exception as e:
            st.error(f"Error reading Word document {file_path}: {e}")
            return ""
    
    def _extract_excel_text(self, file_path: Path) -> str:
        """Extract text from Excel files"""
        try:
            # Try with openpyxl first (for .xlsx)
            if file_path.suffix.lower() == '.xlsx':
                workbook = openpyxl.load_workbook(file_path)
                text = ""
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    text += f"Sheet: {sheet_name}\n"
                    for row in sheet.iter_rows(values_only=True):
                        row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                        if row_text.strip():
                            text += row_text + "\n"
                    text += "\n"
                return text.strip()
            else:
                # Use pandas for .xls files
                df = pd.read_excel(file_path, sheet_name=None)
                text = ""
                for sheet_name, sheet_df in df.items():
                    text += f"Sheet: {sheet_name}\n"
                    text += sheet_df.to_string(index=False) + "\n\n"
                return text.strip()
        except Exception as e:
            st.error(f"Error reading Excel file {file_path}: {e}")
            return ""
    
    def _extract_txt_text(self, file_path: Path) -> str:
        """Extract text from text files"""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                return file.read()
        except UnicodeDecodeError:
            try:
                with open(file_path, 'r', encoding='latin-1') as file:
                    return file.read()
            except Exception as e:
                st.error(f"Error reading text file {file_path}: {e}")
                return ""
        except Exception as e:
            st.error(f"Error reading text file {file_path}: {e}")
            return ""
    
    def _extract_html_text(self, file_path: Path) -> str:
        """Extract text from HTML files"""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                html_content = file.read()
            
            soup = BeautifulSoup(html_content, 'html.parser')
            
            # Remove script and style elements
            for script in soup(["script", "style"]):
                script.decompose()
            
            # Get text and clean it up
            text = soup.get_text()
            
            # Clean up whitespace
            lines = (line.strip() for line in text.splitlines())
            chunks = (phrase.strip() for line in lines for phrase in line.split("  "))
            text = ' '.join(chunk for chunk in chunks if chunk)
            
            return text
            
        except Exception as e:
            st.error(f"Error reading HTML file {file_path}: {e}")
            return ""
    
    def get_file_content_bytes_from_path(self, file_path: str) -> Optional[bytes]:
        """Get file content as bytes for download"""
        try:
            file_path = Path(file_path)
            if file_path.exists():
                with open(file_path, 'rb') as f:
                    return f.read()
            return None
        except Exception as e:
            st.error(f"Error reading file {file_path}: {e}")
            return None

    # NEW: Enhanced US Forms Collection Methods
    def batch_download_us_forms(self, form_urls: List[Dict[str, Any]], progress_callback=None) -> List[Dict[str, Any]]:
        """Batch download US immigration forms with enhanced format detection and validation"""
        downloaded_forms = []
        total_forms = len(form_urls)
        
        st.info(f"Starting batch download of {total_forms} US immigration forms...")
        
        for i, form_info in enumerate(form_urls):
            try:
                if progress_callback:
                    progress_callback(i, total_forms, f"Downloading {form_info.get('form_name', 'Unknown Form')}")
                
                # Enhanced download with format prioritization
                download_result = self.download_us_form_with_format_priority(form_info)
                
                if download_result:
                    downloaded_forms.append(download_result)
                    st.success(f"✅ Downloaded: {form_info.get('form_name', 'Unknown')} ({download_result['file_format']})")
                else:
                    st.warning(f"⚠️ Failed to download: {form_info.get('form_name', 'Unknown')}")
                
                # Brief pause to avoid overwhelming servers
                time.sleep(0.5)
                
            except Exception as e:
                st.error(f"Error downloading {form_info.get('form_name', 'Unknown')}: {e}")
                continue
        
        st.success(f"Batch download completed: {len(downloaded_forms)}/{total_forms} forms successfully downloaded")
        return downloaded_forms
    
    def download_us_form_with_format_priority(self, form_info: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        """Download US form with format prioritization (PDF > DOCX > XLSX > HTML)"""
        base_url = form_info.get('official_source_url', '')
        form_name = form_info.get('form_name', 'Unknown Form')
        
        # Format priority order for US forms
        format_attempts = [
            {'ext': '.pdf', 'format': 'PDF'},
            {'ext': '.docx', 'format': 'DOCX'},
            {'ext': '.xlsx', 'format': 'XLSX'},
            {'ext': '.html', 'format': 'HTML'}
        ]
        
        # Try different format variations
        for format_attempt in format_attempts:
            try:
                # Try direct URL first
                test_url = base_url
                if not test_url.lower().endswith(format_attempt['ext']):
                    # Try appending format extension
                    test_url = base_url.rstrip('/') + format_attempt['ext']
                
                # Validate URL before attempting download
                is_valid, status_code, error_msg = self.validate_url(test_url)
                if is_valid:
                    download_result = self.download_document(test_url, "USA", form_info.get('category', 'Immigration'))
                    if download_result:
                        download_result.update({
                            'form_name': form_name,
                            'form_id': form_info.get('form_id', ''),
                            'category': form_info.get('category', ''),
                            'governing_authority': form_info.get('governing_authority', ''),
                            'discovery_method': 'US Forms Collection',
                            'format_priority_used': format_attempt['format']
                        })
                        return download_result
                
            except Exception as e:
                st.debug(f"Format attempt {format_attempt['format']} failed for {form_name}: {e}")
                continue
        
        # If all format attempts fail, try the original URL
        try:
            download_result = self.download_document(base_url, "USA", form_info.get('category', 'Immigration'))
            if download_result:
                download_result.update({
                    'form_name': form_name,
                    'form_id': form_info.get('form_id', ''),
                    'category': form_info.get('category', ''),
                    'governing_authority': form_info.get('governing_authority', ''),
                    'discovery_method': 'US Forms Collection',
                    'format_priority_used': 'Original URL'
                })
                return download_result
        except Exception as e:
            st.error(f"All download attempts failed for {form_name}: {e}")
        
        return None
    
    def extract_text_with_format_analysis(self, file_path: str) -> Dict[str, Any]:
        """Extract text with enhanced format analysis for US forms"""
        basic_text = self.extract_text(file_path)
        
        file_path_obj = Path(file_path)
        file_extension = file_path_obj.suffix.lower()
        
        analysis_result = {
            'extracted_text': basic_text,
            'text_length': len(basic_text),
            'file_format': file_extension.upper().lstrip('.'),
            'extraction_method': self._get_extraction_method(file_extension),
            'text_quality_score': self._calculate_text_quality_score(basic_text),
            'contains_form_fields': self._detect_form_fields(basic_text),
            'language_detected': 'English',  # Default for US forms
            'extraction_timestamp': time.time()
        }
        
        return analysis_result
    
    def _get_extraction_method(self, file_extension: str) -> str:
        """Get the extraction method used for the file format"""
        method_map = {
            '.pdf': 'PyPDF2',
            '.docx': 'python-docx',
            '.doc': 'python-docx',
            '.xlsx': 'openpyxl',
            '.xls': 'pandas',
            '.txt': 'direct_read',
            '.html': 'BeautifulSoup',
            '.htm': 'BeautifulSoup'
        }
        return method_map.get(file_extension, 'unknown')
    
    def _calculate_text_quality_score(self, text: str) -> float:
        """Calculate a quality score for extracted text (0-1)"""
        if not text or len(text) < 10:
            return 0.0
        
        score = 0.0
        
        # Length factor (longer text generally better, up to a point)
        length_score = min(len(text) / 5000, 1.0) * 0.3
        score += length_score
        
        # Readability factor (presence of common words)
        common_words = ['form', 'application', 'name', 'date', 'address', 'immigration', 'visa', 'petition']
        word_matches = sum(1 for word in common_words if word.lower() in text.lower())
        readability_score = min(word_matches / len(common_words), 1.0) * 0.4
        score += readability_score
        
        # Structure factor (presence of formatting indicators)
        structure_indicators = ['\n', ':', '.', '(', ')', '-', '_']
        structure_count = sum(text.count(indicator) for indicator in structure_indicators)
        structure_score = min(structure_count / 100, 1.0) * 0.3
        score += structure_score
        
        return min(score, 1.0)
    
    def _detect_form_fields(self, text: str) -> bool:
        """Detect if text contains form field indicators"""
        field_indicators = [
            'name:', 'date:', 'address:', 'phone:', 'email:', 'signature:',
            '___', '[ ]', '( )', 'check one:', 'select all that apply:',
            'part i', 'part ii', 'part iii', 'section a', 'section b'
        ]
        
        text_lower = text.lower()
        return any(indicator in text_lower for indicator in field_indicators)
    
    def validate_us_form_completeness(self, downloaded_forms: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Validate completeness of US forms collection"""
        
        # Expected US immigration form categories and their typical forms
        expected_categories = {
            'Family-Based Immigration': {
                'expected_forms': ['I-130', 'I-485', 'I-864', 'I-131', 'I-765'],
                'min_expected': 5
            },
            'Employment-Based Immigration': {
                'expected_forms': ['I-140', 'I-485', 'I-765', 'I-131', 'PERM'],
                'min_expected': 4
            },
            'Naturalization': {
                'expected_forms': ['N-400', 'N-600', 'N-565'],
                'min_expected': 2
            },
            'Temporary Visas': {
                'expected_forms': ['I-129', 'I-539', 'DS-160'],
                'min_expected': 3
            },
            'Asylum and Refugee': {
                'expected_forms': ['I-589', 'I-730', 'I-765'],
                'min_expected': 2
            }
        }
        
        # Analyze collected forms
        collected_by_category = {}
        format_distribution = {}
        total_forms = len(downloaded_forms)
        
        for form in downloaded_forms:
            category = form.get('category', 'Uncategorized')
            form_id = form.get('form_id', 'Unknown')
            file_format = form.get('file_format', 'Unknown')
            
            # Category analysis
            if category not in collected_by_category:
                collected_by_category[category] = []
            collected_by_category[category].append(form_id)
            
            # Format distribution
            format_distribution[file_format] = format_distribution.get(file_format, 0) + 1
        
        # Calculate completeness scores
        category_scores = {}
        overall_completeness = 0
        
        for category, expectations in expected_categories.items():
            collected_forms = collected_by_category.get(category, [])
            expected_forms = expectations['expected_forms']
            min_expected = expectations['min_expected']
            
            # Calculate coverage
            found_expected = [form for form in collected_forms if any(exp in form for exp in expected_forms)]
            coverage_ratio = len(found_expected) / len(expected_forms)
            quantity_ratio = len(collected_forms) / min_expected
            
            category_score = min((coverage_ratio + quantity_ratio) / 2, 1.0)
            category_scores[category] = {
                'score': category_score,
                'collected_count': len(collected_forms),
                'expected_count': min_expected,
                'found_expected_forms': found_expected,
                'status': 'Complete' if category_score >= 0.8 else 'Partial' if category_score >= 0.5 else 'Incomplete'
            }
            
            overall_completeness += category_score
        
        overall_completeness = overall_completeness / len(expected_categories) if expected_categories else 0
        
        # Generate recommendations
        recommendations = []
        
        if overall_completeness < 0.7:
            recommendations.append("Consider expanding search queries to capture more forms")
        
        if format_distribution.get('PDF', 0) < total_forms * 0.5:
            recommendations.append("Prioritize PDF format collection for better text extraction quality")
        
        for category, score_info in category_scores.items():
            if score_info['score'] < 0.6:
                recommendations.append(f"Focus on collecting more forms in {category} category")
        
        if not recommendations:
            recommendations.append("Collection appears comprehensive - consider periodic updates")
        
        return {
            'total_forms_collected': total_forms,
            'overall_completeness_score': overall_completeness * 100,
            'overall_status': 'Excellent' if overall_completeness >= 0.9 else 'Good' if overall_completeness >= 0.7 else 'Needs Improvement',
            'category_analysis': category_scores,
            'format_distribution': format_distribution,
            'recommendations': recommendations,
            'validation_timestamp': time.time()
        }
